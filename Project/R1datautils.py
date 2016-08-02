"""
Written By: Kathryn Cogert
For: Winkler Lab Reactor Data Collection System
Purpose: Update reactor 1 master data file with data collected from probes.
"""

# Import Libraries
import os, time, datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from itertools import islice
from downloader import get_val_from, get_values_from, NotAvailable
from googledriveutils import find_r1masterfile



def remove_file(filename):
    """
    Determines if a file exists before trying to delete it
    :param filename: str, name of file to delete
    :return: boolean if the filename doesn't exist
    """
    if os.path.exists(filename):
        os.remove(filename)
    else:
        return False
    return

# Define constants
COL_LABEL = '\nProbe - '
# TODO: ORP PROBE: REVISE THIS DATE when orp probe is added
IGNORE_BEFORE = pd.to_datetime('5.24.2016')
PROBE_DICT = {'DO (mg/L)': 'DO mg/L',
              'pH': 'pH',
              'NH4+ (mgN/L)': 'NH4 mg/L',
              'ORP (mV)': 'ORP mV'}
TS = '\nTimestamps'
CYCLE_TIMING = ['Begin', 'End of Anaerobic', 'End of Aerobic']
CYCLE_COLUMNS = ['Timestamps', 'NO2-', 'NO3-', 'NH4+']
SODIUM = 22.9898
NITROGEN = 14.0067
OXYGEN = 15.9994
HYDROGEN = 1.00794
CHLORINE = 35.453
MEDIA_VOL = 10

def save_to_workbook(newval,
                     date,
                     header,
                     rows_to_skip=12,
                     wbname='temp.xlsx',
                     sheet_name='Reactor Data'):
    wb = load_workbook(wbname)
    ws = wb[sheet_name]
    for cell in ws[rows_to_skip+1]:
        # TODO: Error if header isn't found
        if cell.value == header:
            colno = cell.col_idx
            break

    for row in ws.iter_rows(min_row=rows_to_skip+1, min_col=1, max_col=1):
        for cell in row:
        # TODO: Error if date isn't found
            if cell.value == date:
                rowno = cell.row
                break

    ws.cell(row=rowno, column=colno).value = newval
    wb.save(wbname)
    return


def dload_r1masterfile(filename):
            master_file = find_r1masterfile()
            try:
                master_file.GetContentFile(filename)
            except Exception, e:
                print "Warning: Something wrong with file R1 Master File."
                print str(e)
                # TODO: add an email alarm to responsible user


def save_r1masterfile(csv,
                      rows_to_skip=12,
                      filename='R1MasterFile.xlsx',
                      sheet_name='Reactor Data',
                      path='.'):
    # TODO: Get this to save in a specified directory rather than locally.
    # TODO: Unit test this.
    if csv: # If we want to save the file, see if it has been saved recently.
        try:
            file_create_date = time.ctime(os.path.getctime('./' + filename))
        except:
            file_create_date = '1.1.1900'
        file_create_date = pd.to_datetime(file_create_date)
        now = datetime.datetime.now()
        if now-file_create_date > datetime.timedelta(days=1):
            dload = True
        else:
            dload = False
    else:
        dload = True
    if dload:
        print 'Downloading master file...'
        dload_r1masterfile(filename)
    # convert to dataframe
    wb = load_workbook(filename, data_only=False)
    # delete the local file if we didn't ask for it.
    if not csv:
        remove_file(filename)
    ws = wb[sheet_name]
    data = ws.values
    data = list(data)[rows_to_skip:]

    cols = list(data[0])
    del cols[0]
    del data[0]
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)
    df.dropna(how='all', inplace=True)
    df.replace('#N/A', np.nan, inplace=True)
    parse_excel = lambda x: eval(str(x)[1:]) if isinstance(x, str) else x
    parse_no3 = lambda x: eval(str(x)[1:x.find('-')]) \
        if isinstance(x, str) else x
    # Parse equation columns that have no references
    for col in df.columns:
        try:
            df[col] = df[col].map(parse_excel)
        except:
            pass
    # Parse nitrate gallery measurements
    for col in df.columns:
        if 'NO3- (mgN/L)\nGallery' in col:
            df[col] = df[col].map(parse_no3)-df[col.replace('3', '2')]

    # Calculate Total Influent Flowrate
    df['Flowrate In (ml/min)\nPump Flowrates'] \
        = df['N Media In (ml/min)\nPump Flowrates'] \
        + df['C Media In (ml/min)\nPump Flowrates'] \
        + df['Water In (ml/min)\nPump Flowrates']

    # Calculate total volume in per cycle
    df['Influent Volume (L)\nOther'] \
        = (df['Plug Flow (min)\nCycle Timing']
        + df['Feed+Aerate (min)\nCycle Timing']) \
        * df['Flowrate In (ml/min)\nPump Flowrates']

    # Calculate NO2- influent concentration
    df['NO2- In (mgN/L)\nCalculated Influent Concentration'] \
        = df['NaNO2 in Media (g)\nOther']/(SODIUM + NITROGEN + OXYGEN*2)\
        * NITROGEN * 1000 / MEDIA_VOL \
        * df['N Media In (ml/min)\nPump Flowrates'] \
        / df['Flowrate In (ml/min)\nPump Flowrates']

    # Calculate NH4+ influent concentration
    df['NH4+ In (mgN/L)\nCalculated Influent Concentration'] \
        = df['NH4Cl in Media (g)\nOther']\
        / (CHLORINE + NITROGEN + HYDROGEN*4)\
        * NITROGEN * 1000 / MEDIA_VOL \
        * df['N Media In (ml/min)\nPump Flowrates'] \
        / df['Flowrate In (ml/min)\nPump Flowrates']

    # Calculate total N removed per cycle
    df['Total N Rem/Cycle (mgN/L) \nCalculations'] \
        = df.filter(regex='Calculated Influent Concentration').sum(axis=1)\
        - df.filter(regex='Gallery - End of Aerobic').sum(axis=1)

    # Calculate NH4+ probe error
    nh4gal = df.filter(regex='NH4\+ \(mgN/L\)\nGallery')
    nh4probe = df.filter(regex='NH4\+ \(mgN/L\)\nProbe')
    nh4gal.columns = CYCLE_TIMING
    nh4probe.columns = CYCLE_TIMING
    df['NH4 Probe Error %\nCalculations'] \
        = ((nh4probe-nh4gal)/nh4gal).mean(axis=1)
    return df
save_r1masterfile(True)

def upload_r1masterfile(filename='temp.xlsx'):
    # Get the file we want
    master_file = find_r1masterfile()
    try:
        master_file.SetContentFile(filename)
        master_file.Upload()
    except Exception, e:
        print "Warning: Something wrong with file R1 Master File."
        print str(e)
        # TODO: add an email alarm to responsible user


def populate_r1masterfile(rows_to_skip=12, filename='temp.xlsx'):
    # Get the R1 master file as a file
    masterdf = save_r1masterfile(True)
    """
    # Convert the juicy stuff to a dataframe
    masterdf = pd.read_excel(filename,
                             sheetname='Reactor Data',
                             encoding="utf-16",
                             skiprows=rows_to_skip,
                             sep='\t',
                             index_col='Date',
                             keep_default_na=False,
                             na_values=['-1.#IND', '1.#QNAN', '1.#IND',
                             '-1.#QNAN', '','N/A', '#NA', 'NA',
                             'NULL', 'NaN', '-NaN', 'nan', '-nan'])
                             """
    # Find what we will populate with probe data
    # Find timestamps
    ts_columns = [col for col in masterdf.columns if TS in col]
    tsdf = masterdf[ts_columns]
    # Find probes, ignore before given date
    probe_columns = [col for col in masterdf.columns if COL_LABEL in col]
    probedf = masterdf[probe_columns]
    probedf = probedf[masterdf.index > IGNORE_BEFORE]
    # Find Indices and column labels of blank values
    stackdf = probedf.stack(dropna=False)
    empty = stackdf[stackdf.isnull()].index.tolist()

    # For each blank look for the probe, time & date of cycle, and return val
    for each in empty:
        probe, time = each[1].split(COL_LABEL)
        time = tsdf.loc[each[0], time+TS]
        ts = each[0]+pd.DateOffset(hour=time.hour, minute=time.minute)
        val = get_val_from(1, ts, PROBE_DICT.get(probe))
        probedf.set_value(each[0], each[1], val)
        # Save that value to the workbook
        save_to_workbook(val, each[0], each[1])
    upload_r1masterfile()
    print 'Master file updated. ' + str(datetime.datetime.now())
    remove_file('temp.xlsx')
    return probedf


def get_cycletiming(date):
    # Set up the dataframe and download the data
    cycledf = pd.DataFrame(index=CYCLE_TIMING,
                                    columns=CYCLE_COLUMNS)
    masterdf = save_r1masterfile(True)
    # Sample Timestamps
    ts = masterdf.filter(regex=TS).loc[date]
    ts.index = CYCLE_TIMING
    for idx, each in enumerate(ts):
            cycledf['Timestamps'].iloc[idx] = pd.to_datetime(date) + \
                                                pd.DateOffset(
                                                    hours=each.hour,
                                                    minutes=each.minute)

    return cycledf, masterdf


def build_cyclemeasuredf(date):
    # Download data and find relevant timestamps
    measuredf, masterdf = get_cycletiming(date)
    # Format Values measured by the gallery
    galvals = masterdf.filter(regex='\nGallery').loc[date]
    measuredf['NO2-'] = galvals.filter(regex='NO2-').values
    measuredf['NO3-'] = galvals.filter(regex='NO3-').values
    measuredf['NH4+'] = galvals.filter(regex='NH4+').values
    return measuredf


def build_cycleprobedf(date, get_blank=False):
    # Download data and find relevant timestamps
    tsdf, masterdf = get_cycletiming(date)
    probe_list = PROBE_DICT.values()  # List of probes we want info for
    data = {each: [None]*3 for each in probe_list}
    probedf = pd.DataFrame(data=data)
    probedf['Date'] = tsdf['Timestamps'].tolist()
    probedf.set_index('Date', inplace=True)
    aerobic_start = tsdf['Timestamps'].iloc[1]
    # If no data is available for that point, say so
    if not get_blank:
        try:
            criodf = get_values_from(1, tsdf['Timestamps'].iloc[0],
                                     timestamp2=tsdf['Timestamps'].iloc[-1])
            aerobic_start = criodf['Gas Pump'].idxmax(0)
            probedf = criodf[probe_list]
            avail = True
        except NotAvailable:
            # If no probe data available, return a blank dataframe.
            avail = False
    return probedf, aerobic_start, avail
