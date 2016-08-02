# sharedpyxl - A small library to allow openpyxl to parse shared formulae
#
# See http://stackoverflow.com/questions/30738584/openpyxl-returning-empty-cell-values-for-formula-series
# and https://bitbucket.org/openpyxl/openpyxl/issue/435/incomplete-read-with-extended-cells
# for background.
#
# Requires Robin Macharg's formula parser (http://www.ewbi.com/ewbi.develop/samples/jsport_nonEAT.html)
# to be importable as xlparse.
#
# Permission is hereby granted, free of charge, to any person obtaining a
# copy of this software and associated documentation files (the "Software"),
# to deal in the Software without restriction, including without limitation
# the rights to use, copy, modify, merge, publish, distribute, sublicense,
# and/or sell copies of the Software, and to permit persons to whom the
# Software is furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
# FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SOFTWARE.

import openpyxl
import os
import imp
import re
from collections import OrderedDict

xlparse = imp.load_source('xlparse', os.getcwd() +
                          '/Project/OpenpyxlAddOns/downloader.py')

class PatchedWs(openpyxl.worksheet.Worksheet):

    "`Worksheet` subclass that preserves the shared-formula read order."

    def __init__(self, parent_workbook, title="Sheet"):
        super(PatchedWs, self).__init__(parent_workbook, title)
        self.formula_attributes = OrderedDict()

openpyxl.reader.worksheet.Worksheet = PatchedWs

def get_formula(cell):
    """
    Retrieve the cell formula/value, allowing for shared formulae.
    `cell`: the `openpyxl.cell.Cell` object, whose formula we want.
    """
    try:
        attrs = cell.parent.formula_attributes[cell.coordinate]
    except KeyError:
        return cell.value
    if attrs.get('t') != 'shared':
        return cell.value
    group_num = attrs['si']
    for coord, nattrs in cell.parent.formula_attributes.items():
        if nattrs.get('t') == 'shared' and nattrs['si'] == group_num:
            master_cell = cell.parent[coord]
            break
    # master_cell is guaranteed to be defined since at minimum our cell
    # will fit the two criteria
    if master_cell is cell:
        return cell.value
    return convert_to_a1(
        convert_to_r1c1(master_cell.value, master_cell.coordinate),
        cell.coordinate)

def convert_to_r1c1(formula, cell_addr):
    """
    Convert a formula to A1 notation.
    `formula`: the R1C1-style formula to convert
    `cell_addr`: the A1-style address of the cell whose formula we are converting
    """
    parser = R1C1Converter(cell_addr)
    parser.parse(formula)
    return parser.render()

def convert_to_a1(formula, cell_addr):
    """
    Convert a formula to R1C1 notation.
    `formula`: the A1-style formula to convert
    `cell_addr`: the A1-style address of the cell whose formula we are converting
    """
    parser = A1Converter(cell_addr)
    parser.parse(formula)
    return parser.render()

class RangeConverter(xlparse.ExcelParser):

    """
    Parser that converts all range reference to either A1 or R1C1 mode.
    `base_cell_addr`: The A1-style address of the cell whose formula
                      we are parsing.
    """

    def __init__(self, base_cell_addr):
        col, row = openpyxl.utils.coordinate_from_string(base_cell_addr)
        self.col = openpyxl.utils.column_index_from_string(col)
        self.row = row

    def parse(self, formula):
        # Extend supermethod to ensure all range tokens are in R1C1 notation
        xlparse.ExcelParser.parse(self, formula)
        for token in self.tokens.items:
            if (token.ttype == self.TOK_TYPE_OPERAND
                and token.tsubtype == self.TOK_SUBTYPE_RANGE):
                token.tvalue = self.convert(token.tvalue)

    def render(self):
        return '=' + xlparse.ExcelParser.render(self)

    def convert(self, addr):
        "Convert the given address to either R1C1 stle or to A1 style."
        if '!' in addr:
            sheet_name, addr = addr.rsplit('!', 1)
            sheet_name = sheet_name + '!'
        else:
            sheet_name = ''
        if ':' in addr:
            lower, upper = addr.split(':')
            return (sheet_name +
                    self.convert_cell(lower) +
                    ':' + self.convert_cell(upper))
        else:
            return sheet_name + self.convert_cell(addr)

    def convert_cell(self, addr):
        "Convert a specific cell's address to A1 or R1C1 notation."
        raise NotImplementedError


class R1C1Converter(RangeConverter):

    "Parser that converts all range references from A1 to R1C1 mode."

    def convert_cell(self, addr):
        # Convert the given cell address from A1 to R1C1 mode
        col, row = openpyxl.utils.coordinate_from_string(addr)
        col = openpyxl.utils.column_index_from_string(col)
        if addr.startswith('$'):  # column-absolute
            colstr = 'C{}'.format(col)
        elif col == self.col:
            colstr = 'C'
        else:
            colstr = 'C[{}]'.format(col - self.col)
        if '$' in addr[1:]:  # row-absolute
            rowstr = 'R{}'.format(row)
        elif row == self.row:
            rowstr = 'R'
        else:
            rowstr = 'R[{}]'.format(row - self.row)
        return rowstr + colstr


class A1Converter(RangeConverter):

    """
    Parser that converts all range references from R1C1 to A1 mode.
    """

    R1C1_RE = re.compile(r'^R(\[[-+]?\d*\]|\d*)C(\[[-+]?\d*\]|\d*)$')

    def convert_cell(self, addr):
        # Convert the given cell address from R1C1 mode to A1
        rowstr, colstr = self.R1C1_RE.match(addr).groups()
        if colstr.startswith('['):  # col-relative
            newcol = self.col + long(colstr[1:-1])
            newcol = openpyxl.utils.get_column_letter(newcol)
        elif not colstr:
            newcol = openpyxl.utils.get_column_letter(self.col)
        else:
            newcol = '$' + openpyxl.utils.get_column_letter(int(colstr))
        if rowstr.startswith('['):  # row-relative
            newrow = str(self.row + long(rowstr[1:-1]))
        elif not rowstr:
            newrow = str(self.row)
        else:
            newrow = '$' + rowstr
        return newcol + newrow
